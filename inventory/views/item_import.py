"""
Views for Excel import/export of items.
"""
import io
from typing import Dict, Any, List, Tuple
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import TemplateView

from .base import InventoryBaseView
from .. import models
from ..forms.base import UNIT_CHOICES


try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ItemExcelTemplateDownloadView(InventoryBaseView, View):
    """View to download Excel template for item import."""
    
    def get(self, request, *args, **kwargs):
        """Generate and download Excel template."""
        if not OPENPYXL_AVAILABLE:
            messages.error(request, _('کتابخانه openpyxl نصب نشده است. لطفاً آن را نصب کنید.'))
            return HttpResponseRedirect(reverse_lazy('inventory:items'))
        
        company_id = request.session.get('active_company_id')
        if not company_id:
            messages.error(request, _('لطفاً ابتدا یک شرکت را انتخاب کنید.'))
            return HttpResponseRedirect(reverse_lazy('inventory:items'))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "کالاها"
        
        # Header row with styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Define columns (convert lazy translations to strings)
        headers = [
            str(_('نوع کالا (کد یا نام)')),
            str(_('دسته‌بندی (کد یا نام)')),
            str(_('زیردسته (کد یا نام)')),
            str(_('کد کاربری (2 رقم)')),
            str(_('نام (فارسی)')),
            str(_('نام (انگلیسی)')),
            str(_('بچ نامبر ثانویه')),
            str(_('قابل فروش (1=بله، 0=خیر)')),
            str(_('رهگیری لات (1=بله، 0=خیر)')),
            str(_('رسید موقت (1=بله، 0=خیر)')),
            str(_('شناسه مالیاتی')),
            str(_('عنوان مالیاتی')),
            str(_('حداقل موجودی')),
            str(_('واحد اصلی')),
            str(_('واحد گزارش')),
            str(_('توضیح کوتاه')),
            str(_('یادداشت‌ها')),
            str(_('ترتیب نمایش')),
            str(_('فعال (1=بله، 0=خیر)')),
            str(_('انبارهای مجاز (کدها با کاما جدا شوند)')),
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=str(header))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add example row
        example_row = [
            '001',  # نوع کالا
            'test',  # دسته‌بندی
            'Test Subcategory',  # زیردسته
            '20',  # کد کاربری
            'نمونه کالا',  # نام فارسی
            'Sample Item',  # نام انگلیسی
            '',  # بچ نامبر ثانویه
            '1',  # قابل فروش
            '0',  # رهگیری لات
            '0',  # رسید موقت
            '',  # شناسه مالیاتی
            '',  # عنوان مالیاتی
            '',  # حداقل موجودی
            'EA',  # واحد اصلی
            'EA',  # واحد گزارش
            '',  # توضیح کوتاه
            '',  # یادداشت‌ها
            '0',  # ترتیب نمایش
            '1',  # فعال
            '001',  # انبارهای مجاز
        ]
        
        for col_num, value in enumerate(example_row, 1):
            ws.cell(row=2, column=col_num, value=value)
        
        # Add data validation and help
        # Get available item types, categories, subcategories, warehouses
        item_types = models.ItemType.objects.filter(company_id=company_id, is_enabled=1).order_by('name')
        categories = models.ItemCategory.objects.filter(company_id=company_id, is_enabled=1).order_by('name')
        subcategories = models.ItemSubcategory.objects.filter(company_id=company_id, is_enabled=1).select_related('category').order_by('category__name', 'name')
        warehouses = models.Warehouse.objects.filter(company_id=company_id, is_enabled=1).order_by('name')
        
        # Add a sheet with reference data
        ref_sheet = wb.create_sheet(title="راهنما")
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 1. انواع کالا
        row = 1
        ref_sheet.cell(row=row, column=1, value=str(_('انواع کالا'))).fill = header_fill
        ref_sheet.cell(row=row, column=1).font = header_font
        ref_sheet.cell(row=row, column=1).alignment = header_alignment
        ref_sheet.merge_cells(f'A{row}:B{row}')
        row += 1
        ref_sheet.append([str(_('کد')), str(_('نام'))])
        for item_type in item_types:
            ref_sheet.append([item_type.public_code, item_type.name])
        row = ref_sheet.max_row + 2
        
        # 2. دسته‌بندی‌ها
        ref_sheet.cell(row=row, column=1, value=str(_('دسته‌بندی‌ها'))).fill = header_fill
        ref_sheet.cell(row=row, column=1).font = header_font
        ref_sheet.cell(row=row, column=1).alignment = header_alignment
        ref_sheet.merge_cells(f'A{row}:B{row}')
        row += 1
        ref_sheet.append([str(_('کد')), str(_('نام'))])
        for category in categories:
            ref_sheet.append([category.public_code, category.name])
        row = ref_sheet.max_row + 2
        
        # 3. زیردسته‌بندی‌ها
        ref_sheet.cell(row=row, column=1, value=str(_('زیردسته‌بندی‌ها'))).fill = header_fill
        ref_sheet.cell(row=row, column=1).font = header_font
        ref_sheet.cell(row=row, column=1).alignment = header_alignment
        ref_sheet.merge_cells(f'A{row}:C{row}')
        row += 1
        ref_sheet.append([str(_('کد')), str(_('نام')), str(_('دسته‌بندی'))])
        for subcategory in subcategories:
            ref_sheet.append([subcategory.public_code, subcategory.name, subcategory.category.name])
        row = ref_sheet.max_row + 2
        
        # 4. انبارها
        ref_sheet.cell(row=row, column=1, value=str(_('انبارها'))).fill = header_fill
        ref_sheet.cell(row=row, column=1).font = header_font
        ref_sheet.cell(row=row, column=1).alignment = header_alignment
        ref_sheet.merge_cells(f'A{row}:B{row}')
        row += 1
        ref_sheet.append([str(_('کد')), str(_('نام'))])
        for warehouse in warehouses:
            ref_sheet.append([warehouse.public_code, warehouse.name])
        row = ref_sheet.max_row + 2
        
        # 5. واحدها
        ref_sheet.cell(row=row, column=1, value=str(_('واحدها'))).fill = header_fill
        ref_sheet.cell(row=row, column=1).font = header_font
        ref_sheet.cell(row=row, column=1).alignment = header_alignment
        ref_sheet.merge_cells(f'A{row}:B{row}')
        row += 1
        ref_sheet.append([str(_('کد')), str(_('نام'))])
        for unit_code, unit_name in UNIT_CHOICES:
            if unit_code:  # Skip empty option
                ref_sheet.append([unit_code, str(unit_name)])
        
        # Adjust column widths for reference sheet
        ref_sheet.column_dimensions['A'].width = 15
        ref_sheet.column_dimensions['B'].width = 30
        ref_sheet.column_dimensions['C'].width = 30
        
        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 20
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="item_import_template.xlsx"'
        
        wb.save(response)
        return response


class ItemExcelImportView(InventoryBaseView, TemplateView):
    """View to handle Excel file upload and import items."""
    template_name = 'inventory/item_import_result.html'
    
    def post(self, request, *args, **kwargs):
        """Process uploaded Excel file."""
        if not OPENPYXL_AVAILABLE:
            messages.error(request, _('کتابخانه openpyxl نصب نشده است. لطفاً آن را نصب کنید.'))
            return HttpResponseRedirect(reverse_lazy('inventory:items'))
        
        company_id = request.session.get('active_company_id')
        if not company_id:
            messages.error(request, _('لطفاً ابتدا یک شرکت را انتخاب کنید.'))
            return HttpResponseRedirect(reverse_lazy('inventory:items'))
        
        if 'excel_file' not in request.FILES:
            messages.error(request, _('لطفاً یک فایل Excel انتخاب کنید.'))
            return HttpResponseRedirect(reverse_lazy('inventory:items'))
        
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, _('فایل باید با فرمت Excel (.xlsx یا .xls) باشد.'))
            return HttpResponseRedirect(reverse_lazy('inventory:items'))
        
        try:
            # Load workbook
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Process rows
            errors = []
            success_count = 0
            duplicate_count = 0
            
            # Get existing items for duplicate check
            existing_items = set(
                models.Item.objects.filter(company_id=company_id).values_list('name', flat=True)
            )
            existing_item_codes = set(
                models.Item.objects.filter(company_id=company_id).exclude(item_code='').values_list('item_code', flat=True)
            )
            
            # Process each row (skip header row)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not any(row):
                    continue
                
                row_errors = []
                
                try:
                    # Parse row data
                    item_data = self._parse_row(row, company_id, row_num)
                    
                    # Validate
                    validation_errors = self._validate_item_data(item_data, company_id, existing_items, existing_item_codes)
                    if validation_errors:
                        row_errors.extend(validation_errors)
                    
                    if row_errors:
                        errors.append({
                            'row': row_num,
                            'errors': row_errors,
                            'data': row[:10]  # First 10 columns for display
                        })
                        continue
                    
                    # Create item
                    item = self._create_item(item_data, company_id, request.user)
                    success_count += 1
                    
                    # Add to existing sets to prevent duplicates in same import
                    existing_items.add(item.name)
                    if item.item_code:
                        existing_item_codes.add(item.item_code)
                    
                except Exception as e:
                    errors.append({
                        'row': row_num,
                        'errors': [str(e)],
                        'data': row[:10] if row else []
                    })
            
            # Prepare context
            context = {
                'success_count': success_count,
                'error_count': len(errors),
                'duplicate_count': duplicate_count,
                'errors': errors,
                'total_rows': ws.max_row - 1,  # Exclude header
            }
            
            if success_count > 0:
                messages.success(request, _('{} کالا با موفقیت وارد شد.').format(success_count))
            if errors:
                messages.warning(request, _('{} ردیف دارای خطا بودند.').format(len(errors)))
            
            return self.render_to_response(context)
            
        except Exception as e:
            messages.error(request, _('خطا در پردازش فایل: {}').format(str(e)))
            return HttpResponseRedirect(reverse_lazy('inventory:items'))
    
    def _parse_row(self, row: Tuple, company_id: int, row_num: int) -> Dict[str, Any]:
        """Parse a row from Excel into item data dictionary."""
        data = {}
        
        # Map columns (0-indexed)
        try:
            data['type_code_or_name'] = str(row[0]).strip() if row[0] else None
            data['category_code_or_name'] = str(row[1]).strip() if row[1] else None
            data['subcategory_code_or_name'] = str(row[2]).strip() if row[2] else None
            data['user_segment'] = str(row[3]).strip() if row[3] else None
            data['name'] = str(row[4]).strip() if row[4] else None
            data['name_en'] = str(row[5]).strip() if row[5] else None
            data['secondary_batch_number'] = str(row[6]).strip() if row[6] else ''
            data['is_sellable'] = self._parse_bool(row[7], default=0)
            data['has_lot_tracking'] = self._parse_bool(row[8], default=0)
            data['requires_temporary_receipt'] = self._parse_bool(row[9], default=0)
            data['tax_id'] = str(row[10]).strip() if row[10] else ''
            data['tax_title'] = str(row[11]).strip() if row[11] else ''
            data['min_stock'] = self._parse_decimal(row[12])
            data['default_unit'] = str(row[13]).strip() if row[13] else None
            data['primary_unit'] = str(row[14]).strip() if row[14] else None
            data['description'] = str(row[15]).strip() if row[15] else ''
            data['notes'] = str(row[16]).strip() if row[16] else ''
            data['sort_order'] = int(row[17]) if row[17] else 0
            data['is_enabled'] = self._parse_bool(row[18], default=1)
            data['warehouse_codes'] = str(row[19]).strip() if row[19] else None
        except (IndexError, ValueError, TypeError) as e:
            raise ValueError(_('خطا در خواندن ردیف: {}').format(str(e)))
        
        return data
    
    def _parse_bool(self, value, default=0) -> int:
        """Parse boolean value from Excel (1/0, yes/no, true/false)."""
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return 1 if value == 1 else 0
        value_str = str(value).strip().lower()
        if value_str in ('1', 'yes', 'true', 'بله', 'y'):
            return 1
        return 0
    
    def _parse_decimal(self, value) -> Decimal:
        """Parse decimal value from Excel."""
        if value is None or value == '':
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    
    def _validate_item_data(self, data: Dict[str, Any], company_id: int, existing_items: set, existing_item_codes: set) -> List[str]:
        """Validate item data and return list of errors."""
        errors = []
        
        # Required fields
        if not data.get('type_code_or_name'):
            errors.append(_('نوع کالا الزامی است.'))
        if not data.get('category_code_or_name'):
            errors.append(_('دسته‌بندی الزامی است.'))
        if not data.get('subcategory_code_or_name'):
            errors.append(_('زیردسته الزامی است.'))
        if not data.get('user_segment'):
            errors.append(_('کد کاربری الزامی است.'))
        elif len(data['user_segment']) != 2 or not data['user_segment'].isdigit():
            errors.append(_('کد کاربری باید دقیقاً 2 رقم عددی باشد.'))
        if not data.get('name'):
            errors.append(_('نام فارسی الزامی است.'))
        if not data.get('name_en'):
            errors.append(_('نام انگلیسی الزامی است.'))
        if not data.get('default_unit'):
            errors.append(_('واحد اصلی الزامی است.'))
        if not data.get('primary_unit'):
            errors.append(_('واحد گزارش الزامی است.'))
        
        # Check duplicates
        if data.get('name') and data['name'] in existing_items:
            errors.append(_('کالایی با این نام قبلاً ثبت شده است.'))
        
        # Validate units
        unit_codes = [code for code, _ in UNIT_CHOICES if code]
        if data.get('default_unit') and data['default_unit'] not in unit_codes:
            errors.append(_('واحد اصلی نامعتبر است.'))
        if data.get('primary_unit') and data['primary_unit'] not in unit_codes:
            errors.append(_('واحد گزارش نامعتبر است.'))
        
        # Validate min_stock
        if data.get('min_stock') is not None and data['min_stock'] < 0:
            errors.append(_('حداقل موجودی نمی‌تواند منفی باشد.'))
        
        return errors
    
    def _create_item(self, data: Dict[str, Any], company_id: int, user) -> models.Item:
        """Create item from validated data."""
        # Resolve type, category, subcategory
        item_type = self._resolve_item_type(data['type_code_or_name'], company_id)
        category = self._resolve_category(data['category_code_or_name'], company_id)
        subcategory = self._resolve_subcategory(
            data['subcategory_code_or_name'], 
            category.id if category else None,
            company_id
        )
        
        if not item_type or not category or not subcategory:
            raise ValueError(_('نوع، دسته‌بندی یا زیردسته یافت نشد.'))
        
        # Resolve warehouses
        warehouses = []
        if data.get('warehouse_codes'):
            warehouse_codes = [code.strip() for code in data['warehouse_codes'].split(',')]
            warehouses = models.Warehouse.objects.filter(
                company_id=company_id,
                public_code__in=warehouse_codes,
                is_enabled=1
            )
        
        # Create item
        item = models.Item(
            company_id=company_id,
            type=item_type,
            category=category,
            subcategory=subcategory,
            user_segment=data['user_segment'],
            name=data['name'],
            name_en=data['name_en'],
            secondary_batch_number=data.get('secondary_batch_number', ''),
            is_sellable=data.get('is_sellable', 0),
            has_lot_tracking=data.get('has_lot_tracking', 0),
            requires_temporary_receipt=data.get('requires_temporary_receipt', 0),
            tax_id=data.get('tax_id', ''),
            tax_title=data.get('tax_title', ''),
            min_stock=data.get('min_stock'),
            default_unit=data['default_unit'],
            primary_unit=data['primary_unit'],
            description=data.get('description', ''),
            notes=data.get('notes', ''),
            sort_order=data.get('sort_order', 0),
            is_enabled=data.get('is_enabled', 1),
            created_by=user,
            edited_by=user,
        )
        
        # Save to generate codes
        item.save()
        
        # Add warehouses
        if warehouses:
            for idx, warehouse in enumerate(warehouses):
                item.warehouses.create(
                    company_id=company_id,
                    warehouse=warehouse,
                    is_primary=1 if idx == 0 else 0,
                )
        
        return item
    
    def _resolve_item_type(self, code_or_name: str, company_id: int):
        """Resolve item type by code or name."""
        if not code_or_name:
            return None
        # Try by code first
        item_type = models.ItemType.objects.filter(
            company_id=company_id,
            public_code=code_or_name,
            is_enabled=1
        ).first()
        if item_type:
            return item_type
        # Try by name
        return models.ItemType.objects.filter(
            company_id=company_id,
            name=code_or_name,
            is_enabled=1
        ).first()
    
    def _resolve_category(self, code_or_name: str, company_id: int):
        """Resolve category by code or name."""
        if not code_or_name:
            return None
        # Try by code first
        category = models.ItemCategory.objects.filter(
            company_id=company_id,
            public_code=code_or_name,
            is_enabled=1
        ).first()
        if category:
            return category
        # Try by name
        return models.ItemCategory.objects.filter(
            company_id=company_id,
            name=code_or_name,
            is_enabled=1
        ).first()
    
    def _resolve_subcategory(self, code_or_name: str, category_id: int, company_id: int):
        """Resolve subcategory by code or name."""
        if not code_or_name:
            return None
        # Try by code first
        subcategory = models.ItemSubcategory.objects.filter(
            company_id=company_id,
            public_code=code_or_name,
            is_enabled=1
        )
        if category_id:
            subcategory = subcategory.filter(category_id=category_id)
        subcategory = subcategory.first()
        if subcategory:
            return subcategory
        # Try by name
        subcategory = models.ItemSubcategory.objects.filter(
            company_id=company_id,
            name=code_or_name,
            is_enabled=1
        )
        if category_id:
            subcategory = subcategory.filter(category_id=category_id)
        return subcategory.first()

